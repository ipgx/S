#!/usr/bin/env python3
"""
Extract segments from all 7 CMS Excel files into standardized JSON format.
Output: {name}_segments.json with fields: id, road, from, to, lat, lon (if available)
"""

import openpyxl
import json
import os
import re

BASE = os.path.dirname(os.path.abspath(__file__))


def clean(val):
    """Clean cell value to string."""
    if val is None:
        return ""
    s = str(val).strip()
    # Remove newlines, excess whitespace
    s = re.sub(r'\s+', ' ', s)
    return s


def extract_apopka():
    """Apopka 2025: Has lat/long. Segment col = 'From to To'."""
    path = os.path.join(BASE, "Apopka 2025.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["2025MasterFile-Stat"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        road = clean(row[1])
        seg_text = clean(row[2])
        lat = row[9]
        lon = row[10]
        if not road or not seg_text:
            continue
        # Parse "Alabama Ave to Sheeler Ave"
        parts = re.split(r'\s+to\s+', seg_text, maxsplit=1, flags=re.IGNORECASE)
        if len(parts) == 2:
            frm, to = parts[0].strip(), parts[1].strip()
        else:
            continue
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        seg = {"id": str(len(segments) + 1), "road": road, "from": frm, "to": to}
        if lat and lon:
            try:
                seg["lat"] = float(lat)
                seg["lon"] = float(lon)
            except (ValueError, TypeError):
                pass
        segments.append(seg)
    wb.close()
    return segments


def extract_hillsborough():
    """Hillsborough 2020: On Street, From, To columns."""
    path = os.path.join(BASE, "Hillsborough 2020.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table 1"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        road = clean(row[0])
        frm = clean(row[1])
        to = clean(row[2])
        if not road or not frm or not to:
            continue
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        segments.append({"id": str(len(segments) + 1), "road": road, "from": frm, "to": to})
    wb.close()
    return segments


def extract_osceola():
    """Osceola 2025: ROADWAY, FROM, TO. Data starts row 5."""
    path = os.path.join(BASE, "Osceola 2025.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["2025 RNCR"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        road = clean(row[2])
        frm = clean(row[3])
        to = clean(row[4])
        if not road or not frm or not to:
            continue
        # Skip header-like rows
        if road.upper() in ("ROADWAY", "FROM", "TO", ""):
            continue
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        segments.append({"id": str(len(segments) + 1), "road": road, "from": frm, "to": to})
    wb.close()
    return segments


def extract_palm_beach():
    """Palm Beach 2024: STN#, ROAD, FROM, TO. Data starts row 3."""
    path = os.path.join(BASE, "Palm Beach County 2024.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table 1"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        stn = row[0]
        road = clean(row[1])
        frm = clean(row[3])
        to = clean(row[5])
        if not road or not frm or not to:
            continue
        if not stn or not str(stn).strip().isdigit():
            continue
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        segments.append({"id": str(stn), "road": road, "from": frm, "to": to})
    wb.close()
    return segments


def extract_polk():
    """Polk 2023: Link, Road Segment, From, To. Data starts row 5. Deduplicate N/S/E/W."""
    path = os.path.join(BASE, "Polk 2023.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table001 (Page 1-11)"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        link = clean(row[0])
        road = clean(row[1])
        frm = clean(row[2])
        to = clean(row[3])
        if not road or not frm or not to:
            continue
        if road.upper() in ("ROAD SEGMENT", "FROM", "TO", ""):
            continue
        # Deduplicate directional pairs (8301N / 8301S -> keep one)
        base_link = re.sub(r'[NSEW]$', '', link)
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        segments.append({"id": base_link, "road": road, "from": frm, "to": to})
    wb.close()
    return segments


def extract_seminole():
    """Seminole 2022: ID, Roadway, Counter Location, Segment Limits. Data starts row 5."""
    path = os.path.join(BASE, "Seminole 2022.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table 1"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=5, values_only=True):
        sid = row[0]
        road = clean(row[1])
        seg_limits = clean(row[3])
        if not road or not seg_limits:
            continue
        if not sid or road.upper() in ("ROADWAY", ""):
            continue
        # Parse "Seminola Blvd to Park Dr"
        parts = re.split(r'\s+to\s+', seg_limits, maxsplit=1, flags=re.IGNORECASE)
        if len(parts) == 2:
            frm, to = parts[0].strip(), parts[1].strip()
        else:
            continue
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        segments.append({"id": str(sid), "road": road, "from": frm, "to": to})
    wb.close()
    return segments


def extract_stlucie():
    """St. Lucie 2024: Roadway Name, Location ('From to To'). Data starts row 4."""
    path = os.path.join(BASE, "StLucie 2024.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Table 1"]
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        road = clean(row[0])
        location = clean(row[1])
        sid = row[2]
        if not road or not location:
            continue
        if road.upper() in ("ROADWAY NAME", ""):
            continue
        # Parse "CITRUS AVE to ORANGE AVE"
        parts = re.split(r'\s+to\s+', location, maxsplit=1, flags=re.IGNORECASE)
        if len(parts) == 2:
            frm, to = parts[0].strip(), parts[1].strip()
        else:
            continue
        key = f"{road}|{frm}|{to}"
        if key in seen:
            continue
        seen.add(key)
        segments.append({"id": str(sid) if sid else str(len(segments) + 1), "road": road, "from": frm, "to": to})
    wb.close()
    return segments


DATASETS = {
    "Apopka": {
        "extract": extract_apopka,
        "region": "Apopka, Orange County, FL",
        "geocode_suffix": "Apopka, FL",
        "county": "Orange",
        "county_fips": "12095",
    },
    "Hillsborough": {
        "extract": extract_hillsborough,
        "region": "Hillsborough County, FL",
        "geocode_suffix": "Hillsborough County, FL",
        "county": "Hillsborough",
        "county_fips": "12057",
    },
    "Osceola": {
        "extract": extract_osceola,
        "region": "Osceola County, FL",
        "geocode_suffix": "Osceola County, FL",
        "county": "Osceola",
        "county_fips": "12097",
    },
    "PalmBeach": {
        "extract": extract_palm_beach,
        "region": "Palm Beach County, FL",
        "geocode_suffix": "Palm Beach County, FL",
        "county": "Palm Beach",
        "county_fips": "12099",
    },
    "Polk": {
        "extract": extract_polk,
        "region": "Polk County, FL",
        "geocode_suffix": "Polk County, FL",
        "county": "Polk",
        "county_fips": "12105",
    },
    "Seminole": {
        "extract": extract_seminole,
        "region": "Seminole County, FL",
        "geocode_suffix": "Seminole County, FL",
        "county": "Seminole",
        "county_fips": "12117",
    },
    "StLucie": {
        "extract": extract_stlucie,
        "region": "St. Lucie County, FL",
        "geocode_suffix": "St. Lucie County, FL",
        "county": "St. Lucie",
        "county_fips": "12111",
    },
}


if __name__ == "__main__":
    print("Extracting segments from all CMS Excel files...\n")
    for name, cfg in DATASETS.items():
        segs = cfg["extract"]()
        out_path = os.path.join(BASE, f"{name}_segments.json")
        with open(out_path, "w") as f:
            json.dump({"name": name, "region": cfg["region"], "segments": segs}, f, indent=2)
        print(f"  {name:15s}: {len(segs):>4} segments -> {out_path}")
        # Show first 3
        for s in segs[:3]:
            print(f"    {s['road']:30s} | {s['from']:25s} -> {s['to']}")
    print("\nDone.")
